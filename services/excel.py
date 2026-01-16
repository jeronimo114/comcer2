import pandas as pd
import os
import time
import datetime
import pypdf
import pandas
import shutil
import pyexcel
import gspread
import gspread_dataframe
import unicodedata
from io import BytesIO
from gspread.utils import ExportFormat
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from logger_config import setup_logger
from oauth2client.service_account import ServiceAccountCredentials
from pathlib import Path
from typing import List, Tuple, Union
from utils import extract_pdf_pages, excel_to_pdf
from services.drive import upload_files


class Client:
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    def __init__(self):
        self.benefit_day = None
        self.logger = setup_logger()
        self.batch: str = None
        self.path = "./assets/base.xlsx"
        self.workbook = load_workbook(self.path)
        self.clients = []
        self.creds = ServiceAccountCredentials.from_json_keyfile_name(
            "credentials.json", self.scopes
        )
        self.generated_files = []
        self.sheets_api_client = gspread.authorize(self.creds)
        self.spreadsheet = self.sheets_api_client.open_by_key(
            "18AMyp1SzQR_3xe7EAH5muotOkNMy7rJV73YVwfExbi8"
        )
        self.consecutivos = []

    def clear_sheet_range(
        self,
        worksheet,
        start_row: int,
        end_row: int,
        start_col: str = "A",
        end_col: str = "Z",
    ):
        range_to_clear = f"{start_col}{start_row}:{end_col}{end_row}"
        worksheet.batch_clear([range_to_clear])

    def fill_info(self, body: dict):
        self.clients = []
        worksheet = self.spreadsheet.get_worksheet(0)
        self.batch = body["batch"]

        # Prepare all values in a single update
        cells_to_update = [
            ("A2", body["batch"]),
            ("B2", body["createdAt"]),
            ("C2", body["register"]["createdAt"]),
            ("E2", body["total"]),
            ("I2", body["customerplant"]["label"]),
            ("J2", body["customerinvoice"]["label"]),
            ("D2", body["disembark"]["createdAt"]),
            ("G2", body["totalweight"]),
            ("H2", body["averageweight"]),
            ("F2", body["individualssumary"]["weigthed"]),
            ("A10", body["batch"]),
            ("B10", body["individualssumary"]["beneficiaries"]),
            ("C10", body["benefitdate"]),
            ("D10", body["databenefit"]["rcc"]),
            ("E10", body["databenefit"]["rcr"]),
            ("F10", body["databenefit"]["pcc"]),
            ("G10", body["databenefit"]["pcr"]),
            ("H10", body["databenefit"]["ml"]),
            ("I10", body["databenefit"]["mckg"]),
            ("J10", body["individualssumary"]["avgbackfat"]),
            ("K10", body["customerplant"]["label"]),
            ("L10", body["customerinvoice"]["label"]),
            ("M10", body["property"]["label"]),
        ]

        # Convert to batch update format
        cell_updates = []
        for cell_addr, value in cells_to_update:
            col = gspread.utils.a1_to_rowcol(cell_addr)[1]
            row = gspread.utils.a1_to_rowcol(cell_addr)[0]
            cell_updates.append({"range": cell_addr, "values": [[value]]})

        # Execute single batch update
        worksheet.batch_update(cell_updates)
        self.logger.info("Filled info sheet successfully")
        self.logger.info("Clearing rows 18-25, columns A and D:J (preserving B:C for formulas)")
        self.clear_sheet_range(worksheet, 18, 25, "A", "A")
        self.clear_sheet_range(worksheet, 18, 25, "D", "J")

        # Prepare all dispatch data in a single batch
        batch_updates = []

        if body["dispatched"]:
            for idx, row in enumerate(body["dispatched"], start=18):

                self.clients.append(row["namedestination"])
                self.logger.info(row["namedestination"])
                # Write batch to column A
                batch_updates.append({"range": f"A{idx}", "values": [[body["batch"]]]})
                # Write remaining data to columns D:J (skip B and C)
                values = [
                    [
                        row["quantityprocessed"],
                        row["quantityvisceras"],
                        0,
                        0,
                        row["namedestination"],
                        body["customerplant"]["label"],
                        body["customerinvoice"]["label"],
                    ]
                ]
                batch_updates.append({"range": f"D{idx}:J{idx}", "values": values})
        else:
            self.clients.append("")

        worksheet.batch_update(batch_updates)
        self.clients = set(self.clients)
        self.get_dispatch_details(body)
        self.get_vehicle_dispatch_dates(body)

    def generate_temp_files(self):
        try:
            os.mkdir(f"./downloads/{self.batch}")
        except FileExistsError:
            shutil.rmtree(f"./downloads/{self.batch}")
            os.mkdir(f"./downloads/{self.batch}")
        for client in self.clients:
            self.logger.info(f"Creating {self.batch}-{client}.xlsx")
            self.logger.info(self.path)
            shutil.copy(
                self.path, f"./downloads/{self.batch}/{self.batch}-{client}.xlsx"
            )
            self.generated_files.append(
                f"./downloads/{self.batch}/{self.batch}-{client}.xlsx"
            )

    def fill_files(self, results_lote, results_individual):
        for client, path in zip(self.clients, self.generated_files):
            workbook = load_workbook(path)

            self.fill_info()
            # 1. fill despacho
            self.fill_despacho(workbook, results_individual, client, path)
            # 2. fill liquidicación
            self.fill_liquidacion(workbook, results_lote, client, path)

            self.fill_sheet(client, path)

    def fill_despacho(self, body: list, client):
        worksheet = self.spreadsheet.get_worksheet(1)
        self.logger.info("Clearing despacho sheet")
        self.clear_sheet_range(worksheet, 2, 90, "A", "R")

        batch = body[0]["batch"].split("-")[1]
        current_row = 2
        # Prepare batch updates for all rows
        batch_updates = []

        for individual in body:
            dest_value = individual["destination"]["value"]
            self.logger.info(dest_value)
            if client != individual["destination"]["label"]:
                continue

            dispatch_info = (
                self.dispatch_details[dest_value]
                if dest_value in self.dispatch_details.keys()
                else None
            )

            plate = dispatch_info["plate"] if dispatch_info else None
            code = dispatch_info["code"] if dispatch_info else None
            load_dates = self.get_load_dates_by_plate(plate) if dispatch_info else None

            # Create row values
            row_values = [
                load_dates[0] if dest_value else 0,  # A
                load_dates[1] if dest_value else 0,  # B
                f"{batch}-{individual['consecutive']}",  # C
                "",  # D (empty)
                individual["property"]["label"],  # E
                individual["ppe"],  # F
                individual["pcc"],  # G
                "",  # H (empty)
                individual["pcr"],  # I
                individual["gd"],  # J
                individual["ml"],  # K
                individual["seurop"],  # L
                individual["mc"],  # M
                individual["mckg"],  # N
                individual["indexpse"],  # O
                individual["destination"]["label"] if dest_value else 0,  # P
                plate if dest_value else 0,  # Q
                code if dest_value else 0,  # R
            ]

            # Add to batch updates
            batch_updates.append(
                {"range": f"A{current_row}:R{current_row}", "values": [row_values]}
            )
            current_row += 1
        self.count = current_row - 18
        worksheet.batch_update(batch_updates)
        self.logger.info(f"Updated second sheet successfully for client {client}")

    def get_consecutivo(self, workbook, path):
        worksheet = workbook["Consec"]

        for row in worksheet.iter_rows(values_only=True):
            self.logger.info(row)

    def get_vehicle_dispatch_dates(self, body: dict):
        # vehicle plates
        vehicles = []

        for dispatch in body["dispatched"]:
            for vehicle_dispatch in dispatch["vehiclesdispatch"]:
                vehicle = {
                    "plate": vehicle_dispatch["plate"],
                    "start_date": vehicle_dispatch["startdate"],
                    "end_date": vehicle_dispatch["enddate"],
                }
                vehicles.append(vehicle)
        self.logger.info("Getting vehicles")
        self.vehicles = vehicles

    def get_load_dates_by_plate(self, plate) -> tuple:
        try:
            if self.vehicles:
                vehicle = list(filter(lambda x: x["plate"] == plate, self.vehicles))[0]
                return (vehicle["start_date"], vehicle["end_date"])
        except Exception as e:
            self.logger.error(e)
            self.logger.error(f"Could not find vehicle for plate {plate}")
            return ("?", "?")

    def get_dispatch_details(self, body: dict) -> dict:
        self.logger.info("Finding dispatch details")
        self.dispatch_details = {
            elem["iddestination"]: {
                "name": elem["namedestination"],
                "plate": elem["dispatchvehicle"]["plate"],
                "code": elem["dispatch"]["code"],
            }
            for elem in body["dispatched"]
        }

    def get_load_dates_by_client(self, client: str) -> tuple:
        """Get load dates for a specific client using their dispatch details"""
        try:
            # Find dispatch details for this client
            client_dispatch = next(
                (
                    dispatch
                    for dispatch in self.dispatch_details.values()
                    if dispatch["name"].strip().upper() == client.strip().upper()
                ),
                None,
            )

            if not client_dispatch:
                self.logger.warning(f"No dispatch found for client: {client}")
                return ("?", "?")

            # Get the plate from client's dispatch
            client_plate = client_dispatch["plate"]

            # Find vehicle info for this plate
            vehicle = next(
                (v for v in self.vehicles if v["plate"] == client_plate), None
            )

            if vehicle:
                self.logger.info(f"Found vehicle dates for client {client}: {vehicle}")
                return (vehicle["start_date"], vehicle["end_date"])
            else:
                self.logger.warning(f"No vehicle found with plate {client_plate}")
                return ("?", "?")

        except Exception as e:
            self.logger.error(f"Error getting load dates: {str(e)}")
            return ("?", "?")

    def fill_liquidacion(self, body: list, client):
        # llegada L4
        # liquidacion L6
        # sacrificio L7
        worksheet = self.spreadsheet.get_worksheet(
            2
        )  # Assuming "LIQUIDACION" is the third worksheet
        self.logger.info(f"Filling register for client {client}")
        start_date, end_date = self.get_load_dates_by_client(client)
        client_dispatch = None
        for dispatch in self.dispatch_details.values():
            if dispatch["name"].strip().upper() == client.strip().upper():
                client_dispatch = dispatch
                break
        # Prepare batch update for liquidacion fields
        batch_updates = [
            {"range": "L4", "values": [[body["register"]["createdAt"]]]},
            {"range": "L5", "values": [[body["weights"][0]["weightdate"]]]},
            {"range": "L6", "values": [[body["databenefit"]["datebenefit"]]]},
            {"range": "O3", "values": [[self.batch.split("-")[1]]]},
            {"range": "L7", "values": [[start_date]]},
            {"range": "L8", "values": [[end_date]]},
        ]

        self.logger.debug(self.vehicles)
        self.logger.debug(self.dispatch_details)
        # Execute batch update
        worksheet.batch_update(batch_updates)

        # Store benefit day for later use
        self.benefit_day = body["databenefit"]["datebenefit"]

    def download_sheet(self, client) -> str:
        """Download the spreadsheet as Excel file"""
        try:
            self.logger.info(
                f"Downloading spreadsheet for lote {self.batch} and client {client}"
            )
            # Get the spreadsheet as bytes
            spreadsheet_data = self.spreadsheet.export(format=ExportFormat.EXCEL)

            # Create downloads directory if it doesn't exist
            download_dir = "downloads"
            if not os.path.exists(download_dir):
                os.makedirs(download_dir)

            # Clean client name: remove spaces and forward slashes
            formatted_client = client.replace(" ", "_").replace("/", "_").strip()

            self.logger.info(f"Formatted client name: {formatted_client}")

            if not os.path.exists(download_dir + f"/{self.batch}"):
                os.makedirs(download_dir + f"/{self.batch}")

            # Save the file with formatted client name
            filename = f"{self.batch}-{formatted_client}.xlsx"
            filepath = os.path.join(download_dir + f"/{self.batch}", filename)

            with open(filepath, "wb") as f:
                f.write(spreadsheet_data)

            self.logger.info(f"Spreadsheet saved to {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(e)
            self.logger.error(f"Error downloading spreadsheet: {str(e)}")
            raise

    def download_consecutivos_sheet(self):
        try:

            download_dir = "downloads/" + self.batch + "/"
            if not os.path.exists(download_dir):
                os.makedirs(download_dir)

            # Connect to destination spreadsheet
            dest_spreadsheet = self.sheets_api_client.open_by_key(
                "12RXnw6ZBzgG4Yn0EvUZbgf2esJ-fFDdL-uEvcTpuS8w"  # Replace with your actual spreadsheet key
            )

            spreadsheet_data = dest_spreadsheet.export(format=ExportFormat.EXCEL)

            # Save the file
            filename = f"Consecutivos.xlsx"
            filepath = os.path.join(download_dir, filename)

            with open(filepath, "wb") as f:
                f.write(spreadsheet_data)

            self.logger.info(f"Spreadsheet saved to {filepath}")
            return filepath
        except Exception as e:
            self.logger.error(e)
            self.logger.error(f"Error downloading spreadsheet: {str(e)}")
            raise

    def copy_consecutivo_row(self, row_number: int):
        """
        Copy a specific row from 'Consec' sheet and upload it to another Google Sheet

        Args:
            row_number: The row number to copy from the Consec sheet
        """
        try:
            # Get source worksheet
            source_worksheet = self.spreadsheet.worksheet("Consec")

            # Get the values from the specified row
            row_values = source_worksheet.row_values(row_number)

            if not any(row_values):  # Skip empty rows
                self.logger.warning(f"Row {row_number} is empty, skipping")
                return

            # Connect to destination spreadsheet
            dest_spreadsheet = self.sheets_api_client.open_by_key(
                "12RXnw6ZBzgG4Yn0EvUZbgf2esJ-fFDdL-uEvcTpuS8w"
            )
            dest_worksheet = dest_spreadsheet.sheet1

            # Get current values to find last row
            existing_values = dest_worksheet.get_all_values()
            next_row = len(existing_values) + 1

            # Sanitize values
            row_values = [self.sanitize_value(value) for value in row_values]

            # Append the row to the next available row
            dest_worksheet.insert_row(
                row_values, next_row, value_input_option="USER_ENTERED"
            )

            self.logger.info(f"Appended row to position {next_row}: {row_values}")

        except gspread.exceptions.APIError as e:
            self.logger.error(f"Google Sheets API error: {str(e)}")
            raise
        except Exception as e:
            self.logger.error(f"Error copying consecutivo row: {str(e)}")
            raise

    def format_benefit_day(self, date_str: str) -> str:
        """Format benefit day from '2025-07-23 01:02:51' to '23_07_2025'"""
        try:
            # Parse the datetime string
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            # Format as DD_MM_YYYY
            return dt.strftime("%d_%m_%Y")
        except Exception as e:
            self.logger.error(f"Error formatting benefit day: {str(e)}")
            return date_str

    def download_sheet_pdf(self, client) -> str:
        """Download a specific sheet as PDF file"""
        try:
            self.logger.info(f"Downloading sheet as PDF")

            # Get the specific sheet or use the first one

            # Get the worksheet as PDF bytes
            spreadsheet_data = self.spreadsheet.export(format=ExportFormat.PDF)

            # Create downloads directory if it doesn't exist
            download_dir = "downloads/" + self.batch
            if not os.path.exists(download_dir):
                os.makedirs(download_dir)

            formatted_date = self.format_benefit_day(self.benefit_day)
            formatted_client = client.replace(" ", "_").replace("/", "_").strip()

            # Save the file with formatted date
            filename = f"{self.batch}_{formatted_client}_.pdf"
            filepath = os.path.join(download_dir, filename)

            with open(filepath, "wb") as f:
                f.write(spreadsheet_data)

            extract_pdf_pages(filepath, download_dir + "/" + filename, [3])
            self.logger.info(f"PDF spreadsheet saved to {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(e)
            self.logger.error(f"Error downloading spreadsheet as PDF: {str(e)}")
            raise

    def sanitize_value(self, value: Union[str, int, float]) -> Union[str, float]:
        """
        Sanitize values before writing to spreadsheet.
        Handles:
        - Numbers (e.g., "15025", "1252")
        - Decimal numbers (e.g., "2119,9", "124,7")
        - Currency values (e.g., "$ 17.361.271")
        - Percentages (e.g., "82,7%")
        - Times (e.g., "1:01:15")
        - Dates (e.g., "2025-07-22 09:49:08")
        """
        try:
            if value is None:
                return ""

            if isinstance(value, (int, float)):
                return value

            # Convert to string and clean whitespace and leading apostrophes
            value = str(value).strip().lstrip("'")

            # Handle empty strings
            if not value:
                return ""

            # Handle date-time format (contains both date and time)
            if len(value.split("-")) == 3 and " " in value:
                return value

            # Handle time format (HH:MM:SS)
            if value.count(":") == 2:
                return value

            # Handle currency values with $ symbol
            if value.startswith("$"):
                cleaned = (
                    value.replace("$", "").replace(".", "").replace(",", "").strip()
                )
                try:
                    return float(cleaned)
                except ValueError:
                    return value

            # Handle percentage values
            if value.endswith("%"):
                cleaned = value.rstrip("%").replace(",", ".")
                try:
                    return float(cleaned) / 100
                except ValueError:
                    return value

            # Handle decimal numbers with comma
            if "," in value:
                cleaned = value.replace(",", ".")
                try:
                    return float(cleaned)
                except ValueError:
                    return value

            # Handle regular numbers
            try:
                # Remove any thousand separators and try to convert
                cleaned = value.replace(".", "")
                if cleaned.isdigit():
                    return float(cleaned)
            except ValueError:
                pass

            # If all else fails, return the original value
            return value

        except Exception as e:
            self.logger.error(f"Error sanitizing value '{value}': {str(e)}")
            return value

    def _normalize_text(self, text: str) -> str:
        """
        Normaliza texto removiendo acentos y convirtiendo a minúsculas.
        Útil para comparar headers que pueden tener acentos (Órgano vs Organo).
        """
        if not text:
            return ""
        # Normalizar usando NFD y remover diacríticos
        normalized = unicodedata.normalize('NFD', str(text))
        without_accents = ''.join(
            char for char in normalized
            if unicodedata.category(char) != 'Mn'
        )
        return without_accents.lower().strip()

    def _parse_cantidades_sheet(self, sheet) -> list:
        """
        Parsea la hoja de cantidades decomisadas.
        Maneja múltiples secciones (CALIDAD, CANALES, etc.)
        """
        cantidades = []
        current_section = None
        in_data_section = False

        for row in sheet.iter_rows(values_only=True):
            # Saltar filas completamente vacías
            if not any(row):
                if in_data_section:
                    in_data_section = False
                continue

            first_cell = row[0]

            if first_cell and isinstance(first_cell, str):
                normalized = self._normalize_text(first_cell)

                # Detectar header de datos
                if normalized == "individuo":
                    in_data_section = True
                    continue

                # Detectar nombre de sección (mayúsculas, sin más columnas significativas)
                if first_cell.isupper() and len(first_cell) > 2:
                    # Verificar que las otras columnas estén vacías o casi vacías
                    other_cols = [c for c in row[1:5] if c]
                    if len(other_cols) <= 1:
                        current_section = first_cell.strip()
                        in_data_section = False
                        continue

            # Procesar filas de datos
            if in_data_section and first_cell:
                cantidades.append({
                    "individuo": str(row[0]) if row[0] else "",
                    "organo": str(row[1]) if len(row) > 1 and row[1] else "",
                    "cantidad": float(row[2]) if len(row) > 2 and row[2] else 0.0,
                    "unidad": str(row[3]) if len(row) > 3 and row[3] else "",
                    "fecha_registro": str(row[4]) if len(row) > 4 and row[4] else "",
                    "seccion": current_section or "GENERAL"
                })

        return cantidades

    def _parse_motivos_sheet(self, sheet) -> list:
        """
        Parsea la hoja de motivos de decomisos.
        """
        motivos = []
        in_data_section = False

        for row in sheet.iter_rows(values_only=True):
            # Saltar filas completamente vacías
            if not any(row):
                if in_data_section:
                    break  # Fin de datos
                continue

            first_cell = row[0]

            # Detectar header
            if first_cell and isinstance(first_cell, str):
                if self._normalize_text(first_cell) == "individuo":
                    in_data_section = True
                    continue

            # Procesar filas de datos
            if in_data_section and first_cell:
                # Decomiso Total puede ser "Si"/"No" o booleano
                decomiso_total = row[3] if len(row) > 3 else False
                if isinstance(decomiso_total, str):
                    decomiso_total = decomiso_total.lower() in ["si", "sí", "yes", "true", "1"]

                motivos.append({
                    "individuo": str(row[0]) if row[0] else "",
                    "organo": str(row[1]) if len(row) > 1 and row[1] else "",
                    "patologia": str(row[2]) if len(row) > 2 and row[2] else "",
                    "decomiso_total": bool(decomiso_total),
                    "fecha_registro": str(row[4]) if len(row) > 4 and row[4] else ""
                })

        return motivos

    def parse_decomisos_excel(self, excel_bytes: bytes) -> dict:
        """
        Parsea el Excel de resumen de despacho y extrae las tablas de decomisos.

        Args:
            excel_bytes: Contenido del archivo Excel en bytes

        Returns:
            dict: {"cantidades": [...], "motivos": [...]}
        """
        result = {
            "cantidades": [],
            "motivos": []
        }

        try:
            # Cargar workbook desde bytes
            workbook = load_workbook(BytesIO(excel_bytes), data_only=True)

            self.logger.info(f"Excel sheets found: {workbook.sheetnames}")

            # Parsear hoja "Cantidades decomisadas"
            cantidades_sheet_name = None
            for name in workbook.sheetnames:
                if "cantidades" in self._normalize_text(name):
                    cantidades_sheet_name = name
                    break

            if cantidades_sheet_name:
                self.logger.info(f"Parsing sheet: {cantidades_sheet_name}")
                sheet = workbook[cantidades_sheet_name]
                result["cantidades"] = self._parse_cantidades_sheet(sheet)
            else:
                self.logger.warning("Sheet 'Cantidades decomisadas' not found")

            # Parsear hoja "Motivos de decomisos"
            motivos_sheet_name = None
            for name in workbook.sheetnames:
                if "motivos" in self._normalize_text(name):
                    motivos_sheet_name = name
                    break

            if motivos_sheet_name:
                self.logger.info(f"Parsing sheet: {motivos_sheet_name}")
                sheet = workbook[motivos_sheet_name]
                result["motivos"] = self._parse_motivos_sheet(sheet)
            else:
                self.logger.warning("Sheet 'Motivos de decomisos' not found")

            self.logger.info(f"Parsed {len(result['cantidades'])} cantidades and {len(result['motivos'])} motivos")

        except Exception as e:
            self.logger.error(f"Error parsing decomisos Excel: {str(e)}")

        return result

    def fill_decomisos(self, decomisos_data: dict):
        """
        Escribe los datos de decomisos a Google Sheets en la hoja "Decomisos".

        La hoja tendrá dos tablas:
        - Tabla 1: Cantidades Decomisadas
        - Tabla 2: Motivos de Decomisos (después de tabla 1 + 2 filas vacías)

        Args:
            decomisos_data: dict con "cantidades" y "motivos"
        """
        try:
            # Acceder a la hoja "Decomisos" existente
            worksheet = self.spreadsheet.worksheet("Decomisos")
            self.logger.info("Accessed 'Decomisos' worksheet")

            # Limpiar contenido existente
            worksheet.clear()
            self.logger.info("Cleared existing content in Decomisos sheet")

            batch_updates = []
            current_row = 1

            # ===== Tabla 1: Cantidades Decomisadas =====
            header_cantidades = [
                "Individuo", "Órgano", "Cantidad", "Unidad", "Fecha Registro", "Sección"
            ]
            batch_updates.append({
                "range": f"A{current_row}:F{current_row}",
                "values": [header_cantidades]
            })
            current_row += 1

            # Datos de cantidades
            for item in decomisos_data.get("cantidades", []):
                row_values = [
                    item.get("individuo", ""),
                    item.get("organo", ""),
                    item.get("cantidad", 0),
                    item.get("unidad", ""),
                    item.get("fecha_registro", ""),
                    item.get("seccion", "")
                ]
                batch_updates.append({
                    "range": f"A{current_row}:F{current_row}",
                    "values": [row_values]
                })
                current_row += 1

            # Espacio entre tablas (2 filas vacías)
            current_row += 2

            # ===== Tabla 2: Motivos de Decomisos =====
            header_motivos = [
                "Individuo", "Órgano", "Patología", "Decomiso Total", "Fecha Registro"
            ]
            batch_updates.append({
                "range": f"A{current_row}:E{current_row}",
                "values": [header_motivos]
            })
            current_row += 1

            # Datos de motivos
            for item in decomisos_data.get("motivos", []):
                row_values = [
                    item.get("individuo", ""),
                    item.get("organo", ""),
                    item.get("patologia", ""),
                    "Sí" if item.get("decomiso_total") else "No",
                    item.get("fecha_registro", "")
                ]
                batch_updates.append({
                    "range": f"A{current_row}:E{current_row}",
                    "values": [row_values]
                })
                current_row += 1

            # Ejecutar batch update
            if batch_updates:
                worksheet.batch_update(batch_updates)

            cantidades_count = len(decomisos_data.get("cantidades", []))
            motivos_count = len(decomisos_data.get("motivos", []))
            self.logger.info(f"Filled Decomisos sheet with {cantidades_count} cantidades and {motivos_count} motivos")

        except gspread.exceptions.WorksheetNotFound:
            self.logger.error("Worksheet 'Decomisos' not found in spreadsheet")
            raise
        except Exception as e:
            self.logger.error(f"Error filling Decomisos sheet: {str(e)}")
            raise
